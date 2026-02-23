#!/usr/bin/env python3
"""
H-1B LCA Disclosure Data Ingestion
Source: DOL OFLC disclosure xlsx (downloaded manually)
Aggregates by SOC code + state → demand signals

Usage: python3 scripts/ingest/h1b-lca.py [path-to-xlsx]
Default: ~/Downloads/LCA_Disclosure_Data_FY2025_Q4.xlsx
"""

import sys, os, json, statistics
from collections import defaultdict
from datetime import datetime

# Supabase via REST API (avoid Node dependency)
import urllib.request
import urllib.parse

# Load env
env = {}
env_path = os.path.join(os.path.dirname(__file__), '../../.env.local')
with open(env_path) as f:
    for line in f:
        line = line.strip()
        if '=' in line and not line.startswith('#'):
            k, v = line.split('=', 1)
            env[k.strip()] = v.strip()

SUPABASE_URL = env['NEXT_PUBLIC_SUPABASE_URL']
SUPABASE_KEY = env['SUPABASE_SERVICE_ROLE_KEY']
FISCAL_YEAR = 'FY2025'

def supabase_request(path, method='GET', data=None, params=None):
    url = f"{SUPABASE_URL}/rest/v1/{path}"
    if params:
        url += '?' + urllib.parse.urlencode(params)
    
    headers = {
        'apikey': SUPABASE_KEY,
        'Authorization': f'Bearer {SUPABASE_KEY}',
        'Content-Type': 'application/json',
        'Prefer': 'resolution=merge-duplicates' if method in ('POST',) else 'return=minimal',
    }
    
    body = json.dumps(data).encode() if data else None
    req = urllib.request.Request(url, data=body, headers=headers, method=method)
    
    try:
        with urllib.request.urlopen(req) as resp:
            return json.loads(resp.read()) if resp.read() else None
    except urllib.error.HTTPError as e:
        error_body = e.read().decode()
        print(f"  ✗ Supabase error: {e.code} {error_body[:200]}")
        return None

def supabase_upsert(table, rows):
    """Upsert via POST with Prefer: resolution=merge-duplicates"""
    url = f"{SUPABASE_URL}/rest/v1/{table}"
    headers = {
        'apikey': SUPABASE_KEY,
        'Authorization': f'Bearer {SUPABASE_KEY}',
        'Content-Type': 'application/json',
        'Prefer': 'resolution=merge-duplicates',
    }
    body = json.dumps(rows).encode()
    req = urllib.request.Request(url, data=body, headers=headers, method='POST')
    try:
        with urllib.request.urlopen(req) as resp:
            return True
    except urllib.error.HTTPError as e:
        print(f"  ✗ Upsert error: {e.code} {e.read().decode()[:200]}")
        return False

def main():
    import openpyxl
    
    default_path = os.path.expanduser('~/Downloads/LCA_Disclosure_Data_FY2025_Q4.xlsx')
    file_path = sys.argv[1] if len(sys.argv) > 1 else default_path
    
    print(f"📄 Loading H-1B LCA file: {file_path}")
    print(f"   Reading with openpyxl (streaming mode)...\n")
    
    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    ws = wb.active
    
    # Get headers
    header_row = next(ws.iter_rows(min_row=1, max_row=1))
    headers = [cell.value for cell in header_row]
    
    # Build column index
    col = {h: i for i, h in enumerate(headers) if h}
    
    print(f"Detected columns: SOC_CODE={col.get('SOC_CODE','?')}, CASE_STATUS={col.get('CASE_STATUS','?')}")
    print(f"  WORKSITE_STATE={col.get('WORKSITE_STATE','?')}, EMPLOYER_NAME={col.get('EMPLOYER_NAME','?')}")
    print(f"  PREVAILING_WAGE={col.get('PREVAILING_WAGE','?')}, WAGE_RATE_OF_PAY_FROM={col.get('WAGE_RATE_OF_PAY_FROM','?')}\n")
    
    # Aggregate
    agg = defaultdict(lambda: {
        'soc_title': None, 'total': 0, 'certified': 0, 'denied': 0, 'withdrawn': 0,
        'prev_wages': [], 'offered_wages': [], 'employers': defaultdict(int), 'metros': defaultdict(int),
    })
    
    row_count = 0
    for row in ws.iter_rows(min_row=2):
        cells = [cell.value for cell in row]
        if len(cells) < max(col.values()) + 1:
            continue
        
        soc = str(cells[col.get('SOC_CODE', 7)] or '').strip()
        state = str(cells[col.get('WORKSITE_STATE', -1)] or '').strip() if 'WORKSITE_STATE' in col else ''
        status = str(cells[col.get('CASE_STATUS', 1)] or '').strip().upper()
        soc_title = str(cells[col.get('SOC_TITLE', 8)] or '').strip()
        employer = str(cells[col.get('EMPLOYER_NAME', 19)] or '').strip()
        metro = str(cells[col.get('WORKSITE_COUNTY', -1)] or '').strip() if 'WORKSITE_COUNTY' in col else ''
        
        if not soc or len(soc) < 5 or not state:
            continue
        
        # Normalize SOC
        norm_soc = soc.replace('.00', '')
        key = f"{norm_soc}|{state}"
        
        entry = agg[key]
        if not entry['soc_title']:
            entry['soc_title'] = soc_title
        entry['total'] += 1
        
        if 'CERTIFIED' in status:
            entry['certified'] += 1
        elif 'DENIED' in status:
            entry['denied'] += 1
        elif 'WITHDRAWN' in status:
            entry['withdrawn'] += 1
        
        # Wages
        try:
            pw = float(str(cells[col.get('PREVAILING_WAGE', -1)] or '0').replace(',', '').replace('$', ''))
            if pw > 0:
                entry['prev_wages'].append(pw * 2080 if pw < 500 else pw)
        except (ValueError, TypeError):
            pass
        
        try:
            ow = float(str(cells[col.get('WAGE_RATE_OF_PAY_FROM', -1)] or '0').replace(',', '').replace('$', ''))
            if ow > 0:
                entry['offered_wages'].append(ow * 2080 if ow < 500 else ow)
        except (ValueError, TypeError):
            pass
        
        if employer:
            entry['employers'][employer] += 1
        if metro:
            entry['metros'][metro] += 1
        
        row_count += 1
        if row_count % 100000 == 0:
            print(f"  Processed {row_count:,} rows...")
    
    wb.close()
    print(f"\n📊 {row_count:,} LCA records processed")
    print(f"📈 {len(agg):,} SOC × State combinations\n")
    
    def median_val(arr):
        if not arr: return None
        return round(statistics.median(arr))
    
    def top_n(d, n=5):
        return [k for k, v in sorted(d.items(), key=lambda x: -x[1])[:n]]
    
    # Build state-level rows (3+ applications)
    state_rows = []
    for key, e in agg.items():
        if e['total'] < 3:
            continue
        soc, state = key.split('|')
        state_rows.append({
            'soc_code': soc,
            'soc_title': e['soc_title'],
            'state': state,
            'fiscal_year': FISCAL_YEAR,
            'applications_total': e['total'],
            'applications_certified': e['certified'],
            'applications_denied': e['denied'],
            'applications_withdrawn': e['withdrawn'],
            'median_prevailing_wage': median_val(e['prev_wages']),
            'median_offered_wage': median_val(e['offered_wages']),
            'top_employers': top_n(e['employers']),
            'top_metro_areas': top_n(e['metros']),
            'source': 'dol_lca',
        })
    
    # Build national aggregates
    nat = defaultdict(lambda: {
        'soc_title': None, 'total': 0, 'certified': 0, 'denied': 0, 'withdrawn': 0,
        'prev_wages': [], 'offered_wages': [], 'employers': defaultdict(int), 'metros': defaultdict(int),
    })
    
    for key, e in agg.items():
        soc = key.split('|')[0]
        n = nat[soc]
        if not n['soc_title']:
            n['soc_title'] = e['soc_title']
        n['total'] += e['total']
        n['certified'] += e['certified']
        n['denied'] += e['denied']
        n['withdrawn'] += e['withdrawn']
        n['prev_wages'].extend(e['prev_wages'])
        n['offered_wages'].extend(e['offered_wages'])
        for k, v in e['employers'].items():
            n['employers'][k] += v
        for k, v in e['metros'].items():
            n['metros'][k] += v
    
    nat_rows = []
    for soc, e in nat.items():
        if e['total'] < 5:
            continue
        nat_rows.append({
            'soc_code': soc,
            'soc_title': e['soc_title'],
            'state': 'US',
            'fiscal_year': FISCAL_YEAR,
            'applications_total': e['total'],
            'applications_certified': e['certified'],
            'applications_denied': e['denied'],
            'applications_withdrawn': e['withdrawn'],
            'median_prevailing_wage': median_val(e['prev_wages']),
            'median_offered_wage': median_val(e['offered_wages']),
            'top_employers': top_n(e['employers']),
            'top_metro_areas': top_n(e['metros']),
            'source': 'dol_lca',
        })
    
    all_rows = state_rows + nat_rows
    print(f"📤 Inserting {len(all_rows):,} records ({len(state_rows)} state + {len(nat_rows)} national)...\n")
    
    inserted = 0
    for i in range(0, len(all_rows), 100):
        chunk = all_rows[i:i+100]
        if supabase_upsert('intel_h1b_demand', chunk):
            inserted += len(chunk)
        if (i // 100) % 10 == 0:
            print(f"  {inserted:,} / {len(all_rows):,}")
    
    # Update freshness
    supabase_request('intel_data_freshness', method='PATCH', 
        params={'table_name': 'eq.intel_h1b_demand'},
        data={
            'data_period': f'{FISCAL_YEAR} Q4',
            'data_release_date': '2025-02-11',
            'next_expected_release': 'May 2025 (FY2026 Q1)',
            'records_loaded': inserted,
            'last_refreshed_at': datetime.now().isoformat(),
            'refreshed_by': 'matt',
            'refresh_method': 'manual_import',
            'citation_text': f'U.S. Department of Labor, OFLC LCA Disclosure Data, {FISCAL_YEAR} Q4',
            'citation_url': 'https://www.dol.gov/agencies/eta/foreign-labor/performance',
            'coverage_notes': f'{row_count:,} individual LCA applications aggregated to {len(all_rows)} SOC × state demand signals',
            'known_limitations': 'LCA applications ≠ actual H-1B visas granted; includes renewals/amendments; aggregated to 3+ per SOC/state',
            'is_stale': False,
            'stale_reason': None,
        })
    
    print(f"\n🎯 Done! {inserted:,} H-1B demand records inserted\n")
    print("Top 10 occupations by H-1B demand (national):")
    for i, r in enumerate(sorted(nat_rows, key=lambda x: -x['applications_total'])[:10]):
        print(f"  {i+1}. {r['soc_title'] or r['soc_code']} — {r['applications_total']:,} applications (median: ${r['median_offered_wage'] or 0:,})")

if __name__ == '__main__':
    main()
